import streamlit as st
import os
import tempfile
import whisper
import srt
import datetime
import re
import pandas as pd
from pydub import AudioSegment
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill

st.set_page_config(page_title="Transcript Generator", layout="centered")
st.title("🎬 Video Transcript + Question Timestamps")
st.write("Upload video file(s) directly. We'll detect question changes and generate timestamped Excel + SRT.")

# ─────────────────────────────────────────────────────────────
# SETTINGS
# ─────────────────────────────────────────────────────────────
OFFSET_SECONDS = 1     # Add 4 seconds to every detected timestamp
START_OFFSET   = 10     # First question always starts at this many seconds
TRANSCRIPT_WINDOW = 3   # Collect transcript text within 3s of detected moment

# ─────────────────────────────────────────────────────────────
# QUESTION CHANGE REGEX PATTERNS  (priority: top → bottom)
# ─────────────────────────────────────────────────────────────
QUESTION_PATTERNS = [
    # 1. "moving to/with next question" / "moving on to next question"
    r"\bmoving\s+(to|with|on\s+to)\s+(the\s+)?next\s+question\b",

    # 2. "let's come to question 3" / "let's move to question 5"
    r"\blet[''s]*\s+(come|move)\s+to\s+(question|problem)(\s*(number\s*)?\d+)?\b",

    # 3. "next question" (bare)
    r"\bnext\s+question\b",

    # 4. "question number 3" / "question 3"
    r"\bquestion\s*(number\s*)?\d+\b",

    # 5. "q3" / "q.3"
    r"\bq\.?\s*\d+\b",

    # 6. "problem number 2" / "problem 2"
    r"\bproblem\s*(number\s*)?\d+\b",
]

COMPILED_PATTERNS = [re.compile(p, re.IGNORECASE | re.MULTILINE) for p in QUESTION_PATTERNS]


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def format_ts(seconds):
    """Seconds (float) → MM:SS or HH:MM:SS string"""
    seconds = max(0, int(seconds))
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    if h > 0:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{m:02d}:{s:02d}"


def get_transcript_at(segments, target_sec):
    """Return transcript text from segments within TRANSCRIPT_WINDOW seconds of target_sec."""
    texts = []
    for seg in segments:
        if abs(seg["start"] - target_sec) <= TRANSCRIPT_WINDOW:
            texts.append(seg["text"].strip())
    return " ".join(texts) if texts else ""


def detect_question_changes(segments, use_offset=True):
    """
    Returns list of dicts:
      raw_sec   → original detected second
      adj_sec   → raw_sec + OFFSET_SECONDS
      timestamp → formatted adj_sec
      transcript→ text at that moment
    Also injects a Q1 entry at START_OFFSET seconds.
    """
    hits = []
    seen_secs = set()

    for seg in segments:
        text = seg["text"].strip()
        for pattern in COMPILED_PATTERNS:
            if pattern.search(text):
                raw = seg["start"]
                # De-duplicate: skip if within 5s of a previous hit
                if any(abs(raw - s) < 5 for s in seen_secs):
                    break
                seen_secs.add(raw)
                adj = raw + OFFSET_SECONDS
                hits.append({
                    "raw_sec":    raw,
                    "adj_sec":    adj,
                    "timestamp":  format_ts(adj),
                    "transcript": text,
                })
                break

    # Sort by time
    hits.sort(key=lambda x: x["adj_sec"])

    # Inject first question entry based on filename pattern
    # use_offset=True → 10-12 sec start (file has Q1/Q46/Q91/Q136 pattern)
    # use_offset=False → 00:00 start
    first_sec = START_OFFSET if use_offset else 0
    q_first = {
        "raw_sec":    0,
        "adj_sec":    first_sec,
        "timestamp":  format_ts(first_sec),
        "transcript": get_transcript_at(segments, first_sec),
    }

    # Remove any auto-detected hit too close to first question start
    hits = [h for h in hits if h["adj_sec"] > first_sec + 5]
    hits = [q_first] + hits

    return hits


def style_ws(ws, col_widths):
    """Apply header style + column widths to a worksheet."""
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for i, cell in enumerate(ws[1]):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        if i < len(col_widths):
            ws.column_dimensions[cell.column_letter].width = col_widths[i]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", wrap_text=True)


def make_excel(rows):
    """
    rows: list of dicts with Timestamp, Question No., Transcript
    Returns BytesIO Excel.
    """
    df = pd.DataFrame(rows, columns=["Timestamp", "Question No.", "Transcript"])
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Question Timestamps", index=False)
        ws = writer.sheets["Question Timestamps"]
        style_ws(ws, [15, 15, 80])
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────
uploaded_videos = st.file_uploader(
    "🎥 Upload video file(s) — MP4 / MKV / AVI",
    type=["mp4", "mkv", "avi"],
    accept_multiple_files=True,
)

if uploaded_videos:
    st.success(f"✅ {len(uploaded_videos)} video(s) uploaded. Starting transcription...")

    with tempfile.TemporaryDirectory() as temp_dir:

        model = whisper.load_model("base")
        all_combined_rows = []

        for uploaded_file in uploaded_videos:
            st.markdown("---")
            video_name = uploaded_file.name
            st.subheader(f"🎥 {video_name}")

            # Save uploaded file to disk
            video_path = os.path.join(temp_dir, video_name)
            with open(video_path, "wb") as f:
                f.write(uploaded_file.read())

            st.info("🔁 Extracting audio and transcribing...")

            # Extract audio
            audio_path = os.path.join(temp_dir, "temp_audio.wav")
            try:
                audio = AudioSegment.from_file(video_path)
                audio.export(audio_path, format="wav")
            except Exception as e:
                st.error(f"⚠️ Failed to extract audio: {e}")
                continue

            # Transcribe
            result   = model.transcribe(audio_path)
            segments = result["segments"]

            # ── SRT ──────────────────────────────────────────
            subtitles = []
            for i, seg in enumerate(segments):
                start = datetime.timedelta(seconds=int(seg["start"]))
                end   = datetime.timedelta(seconds=int(seg["end"]))
                subtitles.append(srt.Subtitle(
                    index=i + 1, start=start, end=end,
                    content=seg["text"].strip()
                ))
            srt_text = srt.compose(subtitles)

            with st.expander("📄 View Full Transcript (SRT)"):
                st.text(srt_text)

            # SRT download
            srt_filename = os.path.splitext(video_name)[0] + ".srt"
            srt_path = os.path.join(temp_dir, srt_filename)
            with open(srt_path, "w", encoding="utf-8") as f:
                f.write(srt_text)
            with open(srt_path, "rb") as f:
                st.download_button(
                    label=f"⬇️ Download SRT — {video_name}",
                    data=f,
                    file_name=srt_filename,
                    mime="text/plain",
                    key=f"srt_{video_name}",
                )

            # ── QUESTION START NUMBER FROM FILENAME ──────────
            # Supports: "q5-10.mp4", "5_10.mp4", "Q5to10.mp4", "batch_5.mp4"
            base_name = os.path.splitext(video_name)[0]
            q_start = 1  # default

            range_match = re.findall(r'(\d+)\s*[-_]\s*(?:to\s*)?(\d+)', base_name, re.IGNORECASE)
            if range_match:
                q_start = int(range_match[-1][0])
            else:
                nums = re.findall(r'\d+', base_name)
                if nums:
                    q_start = int(nums[-1])

            # ── DETECT IF FILE HAS Q1/Q46/Q91/Q136 PATTERN ───
            # Pattern: Q<number> at start or after separator e.g. "Q1", "Q46", "Q91", "Q136"
            OFFSET_Q_VALUES = {1, 46, 91, 136}
            q_offset_match = re.findall(r'(?:^|[_\-\s])q(\d+)(?:[_\-\s]|$)', base_name, re.IGNORECASE)
            use_offset = any(int(n) in OFFSET_Q_VALUES for n in q_offset_match) if q_offset_match else False

            if use_offset:
                st.caption(f"📌 Q{q_start} — starts at ~10 sec (Q-series file detected: `{base_name}`)")
            else:
                st.caption(f"📌 Q{q_start} — starts at 00:00 (no Q-series pattern in filename: `{base_name}`)")

            # ── QUESTION TIMESTAMPS ───────────────────────────
            hits = detect_question_changes(segments, use_offset=use_offset)

            st.markdown("#### 🔍 Question Change Timestamps")

            if hits:
                rows = [
                    {
                        "Timestamp":    h["timestamp"],
                        "Question No.": q_start + i,
                        "Transcript":   h["transcript"],
                    }
                    for i, h in enumerate(hits)
                ]

                # Screen preview
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

                # Excel download
                excel_buf = make_excel(rows)
                st.download_button(
                    label=f"📊 Download Excel — {video_name}",
                    data=excel_buf,
                    file_name=os.path.splitext(video_name)[0] + "_questions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"xl_{video_name}",
                )

                # Collect for combined
                for row in rows:
                    all_combined_rows.append({
                        "Video":        os.path.splitext(video_name)[0],
                        "Timestamp":    row["Timestamp"],
                        "Question No.": row["Question No."],
                        "Transcript":   row["Transcript"],
                    })

            else:
                st.warning("⚠️ No question-change phrases detected in this video.")
                st.caption("Tip: Educator should say phrases like 'next question', 'aage badhate hai', 'question no 3' etc.")

        # ── COMBINED EXCEL (when multiple videos) ────────────
        if len(uploaded_videos) > 1 and all_combined_rows:
            st.markdown("---")
            st.markdown("### 📦 Combined Excel — All Videos")

            df_combined  = pd.DataFrame(all_combined_rows)
            buf_combined = BytesIO()

            with pd.ExcelWriter(buf_combined, engine="openpyxl") as writer:
                for vid_name, grp in df_combined.groupby("Video"):
                    sheet_rows = [
                        {
                            "Timestamp":    r["Timestamp"],
                            "Question No.": r["Question No."],
                            "Transcript":   r["Transcript"],
                        }
                        for _, r in grp.iterrows()
                    ]
                    sheet_df   = pd.DataFrame(sheet_rows)
                    sheet_name = vid_name[:31]
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    style_ws(writer.sheets[sheet_name], [15, 15, 80])

            buf_combined.seek(0)
            st.download_button(
                label="📊 Download Combined Excel (All Videos)",
                data=buf_combined,
                file_name="all_videos_questions.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="combined_excel",
            )
